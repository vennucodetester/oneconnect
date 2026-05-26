"""
OneConnect StoreConnect Pulse - Asset Dashboard Scraper
Collects temperature data and screenshots from asset dashboards.
Phase 1: 2 test assets.
"""

import json
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path

import numpy as np
from openpyxl import Workbook
from playwright.sync_api import sync_playwright

# --- Configuration ---
BASE_URL = "https://mc.us.oneconnect.net"
API_BASE = "https://api.us.oneconnect.net/oneconnect-api"
TENANT_ID = 37
AUTH_STATE_FILE = Path(__file__).parent / "auth_state.json"
DATA_DIR = Path(__file__).parent / "data"

# How many assets to scrape (set high enough to cover all ~639 assets)
NUM_ASSETS_TO_SCRAPE = 1000

# How long to wait for dashboard graphs to load (seconds)
GRAPH_LOAD_TIMEOUT = 30


def setup_daily_folder():
    """Create today's output folders."""
    today = datetime.now().strftime("%Y-%m-%d")
    day_dir = DATA_DIR / today
    (day_dir / "screenshots").mkdir(parents=True, exist_ok=True)
    (day_dir / "timeseries").mkdir(parents=True, exist_ok=True)
    return day_dir


def login_and_save_session(playwright):
    """Open browser for manual login, wait for login to complete automatically."""
    print("\n=== First-time login ===")
    print("A browser window will open. Please log in to StoreConnect Pulse.")
    print("The script will detect when you're logged in and continue automatically.\n")

    browser = playwright.chromium.launch(headless=False, slow_mo=500)
    context = browser.new_context(viewport={"width": 1536, "height": 900})
    page = context.new_page()
    page.goto(f"{BASE_URL}/login")

    # Wait until we have a TOKEN in localStorage (max 5 minutes)
    print("Waiting for you to log in...")
    for _ in range(300):  # 5 minutes max
        time.sleep(1)
        try:
            token = page.evaluate("localStorage.getItem('TOKEN')")
            if token and token.startswith("Bearer"):
                break
        except Exception:
            pass
        # Also check URL as backup
        if "/login" not in page.url and "oneconnect.net" in page.url:
            time.sleep(3)
            try:
                token = page.evaluate("localStorage.getItem('TOKEN')")
                if token:
                    break
            except Exception:
                pass
    else:
        print("ERROR: Login timed out after 5 minutes.")
        browser.close()
        return

    # Wait a bit more for the page to fully load
    time.sleep(5)
    print(f"Login detected! Current URL: {page.url}")

    # Save the session
    context.storage_state(path=str(AUTH_STATE_FILE))
    print(f"Session saved to {AUTH_STATE_FILE}")

    browser.close()


def get_asset_list(page, token):
    """Fetch the list of all assets from the API."""
    response = page.evaluate(f"""
        fetch('{API_BASE}/tenants/{TENANT_ID}/tenant-api/asset-optimizer/v1/assets?pageNumber=0&pageSize=1000', {{
            headers: {{'Authorization': '{token}'}}
        }}).then(r => r.text())
    """)
    if response:
        return json.loads(response)
    return None


def get_asset_detail(page, asset_id, token):
    """Fetch detailed info for a single asset."""
    response = page.evaluate(f"""
        fetch('{API_BASE}/tenants/{TENANT_ID}/tenant-api/asset-optimizer/v1/assets/{asset_id}', {{
            headers: {{'Authorization': '{token}'}}
        }}).then(r => r.text())
    """)
    if response:
        return json.loads(response)
    return None


def extract_dashboard_data(page):
    """Read the current dashboard values from the DOM."""
    data = {}

    # Extract all text content from the dashboard area
    # The dashboard cards show: label + value + timestamp
    cards = page.evaluate("""
        (() => {
            const result = {};
            // Look for the value cards at the top of the dashboard
            const allText = document.body.innerText;

            // Control Setpoint
            const setpointMatch = allText.match(/Control Setpoint\\s*([\\d.]+)\\s*°F/);
            if (setpointMatch) result.control_setpoint = parseFloat(setpointMatch[1]);

            // Control Temperature
            const ctrlTempMatch = allText.match(/Control Temperature\\s*([\\d.]+)\\s*°F/);
            if (ctrlTempMatch) result.control_temp = parseFloat(ctrlTempMatch[1]);

            // Defrost Terminate
            const defrostMatch = allText.match(/Defrost Terminate\\s*([\\d.]+)\\s*°F/);
            if (defrostMatch) result.defrost_terminate = parseFloat(defrostMatch[1]);

            // Compressor Line Temperature
            const compMatch = allText.match(/Compressor Line Temperature\\s*([\\d.]+)\\s*°F/);
            if (compMatch) result.compressor_line_temp = parseFloat(compMatch[1]);

            // Status tiles - look for known status names and check if they appear
            // Green = ON/OK, Gray = OFF
            const statusElements = document.querySelectorAll('[class*="widget"], [class*="status"], [class*="tile"]');
            // We'll parse these from text as a fallback
            result.raw_text = allText;

            // Timestamp
            const tsMatch = allText.match(/Last data (\\d{4}-\\d{2}-\\d{2} \\d{2}:\\d{2} [AP]M)/);
            if (tsMatch) result.last_data_timestamp = tsMatch[1];

            return result;
        })()
    """)
    return cards


def extract_status_tiles(page):
    """Extract ON/OFF status of the 7 status tiles from the page."""
    statuses = page.evaluate("""
        (() => {
            const result = {};
            const tileNames = [
                'Evaporator Fan', 'Compressor Status', 'Defrost Status',
                'Light Status', 'Alarm', 'Low Temperature Alarm', 'High Temperature Alarm'
            ];
            // Find elements containing these names and check their color/style
            const allElements = document.querySelectorAll('div, span, p');
            for (const el of allElements) {
                const text = el.textContent.trim();
                for (const name of tileNames) {
                    if (text === name) {
                        // Check the parent/grandparent for background color
                        let parent = el.closest('[style*="background"], [class*="green"], [class*="gray"], [class*="widget"]') || el.parentElement;
                        if (parent) {
                            const style = window.getComputedStyle(parent);
                            const bg = style.backgroundColor;
                            // Green-ish = ON/OK, Gray-ish = OFF
                            // rgb values: green ~(76,175,80) or similar, gray ~(158,158,158)
                            const match = bg.match(/rgb\\((\\d+),\\s*(\\d+),\\s*(\\d+)/);
                            if (match) {
                                const [_, r, g, b] = match.map(Number);
                                result[name] = g > r && g > 100 ? 'ON' : 'OFF';
                            } else {
                                result[name] = 'UNKNOWN';
                            }
                        }
                        break;
                    }
                }
            }
            return result;
        })()
    """)
    return statuses


def get_module_options(page):
    """Get the list of module options from the Asset dropdown."""
    # Find and click the combobox dropdown
    try:
        dropdown = page.locator('role=combobox').first
        dropdown.click()
        time.sleep(2)
    except Exception:
        print("  WARNING: Could not find module dropdown")
        return ["Asset"]

    # Take a snapshot of visible text to find module names
    options = page.evaluate("""
        (() => {
            const options = [];
            // Look for listbox items, option elements, or any dropdown-like container
            const selectors = [
                '[role="option"]',
                '[role="listbox"] > *',
                '[class*="option"]',
                '[class*="menu"] [class*="item"]',
                '[class*="dropdown"] [class*="item"]',
                '[class*="list"] [class*="item"]'
            ];

            for (const sel of selectors) {
                const items = document.querySelectorAll(sel);
                for (const item of items) {
                    // Get first line of text (module name), strip "ID n/a" or "ID xxx"
                    const lines = item.textContent.trim().split('\\n').map(l => l.trim()).filter(Boolean);
                    let name = lines[0];
                    // Remove trailing "ID ..." text that gets concatenated
                    name = name.replace(/ID\\s*.*/i, '').trim();
                    if (name && name.length < 30 && name !== 'Asset' && name.length > 0) {
                        options.push(name);
                    }
                }
                if (options.length > 0) break;
            }

            return [...new Set(options)];
        })()
    """)

    # Close dropdown
    page.keyboard.press("Escape")
    time.sleep(0.5)

    if options:
        return options

    # Fallback: check page text for known module names
    text = page.inner_text("body")
    known = ["Left", "Center/Right", "Center/Left", "Center", "Right"]
    found = [n for n in known if n in text]
    return found if found else ["Asset"]


def select_module(page, module_name):
    """Select a specific module from the Asset dropdown."""
    try:
        # Open dropdown
        dropdown = page.locator('role=combobox').first
        dropdown.click()
        time.sleep(2)

        # Click the option — use partial text match since DOM text may include "ID n/a"
        # Try exact first, then partial
        option = page.locator(f'text=/{module_name}/i').first
        option.click()
        time.sleep(8)  # Wait for dashboard to fully update and graphs to render
    except Exception as e:
        print(f"  WARNING: Could not select module '{module_name}': {e}")


def extract_timeseries_from_charts(page):
    """
    Extract 24hr time-series data from Recharts React fiber tree.
    The portal uses Recharts which stores widgetData in component props.
    Returns a list of dicts with: timestamp, control_temp, defrost_terminate,
    comp_discharge_temp, defrost_status, refrigeration_do (compressor on/off).
    """
    raw_charts = page.evaluate("""
        (() => {
            const results = [];
            const rechartsWrappers = document.querySelectorAll('.recharts-wrapper');

            const getReactFiber = (el) => {
                const key = Object.keys(el).find(k =>
                    k.startsWith('__reactInternalInstance') ||
                    k.startsWith('__reactFiber')
                );
                return key ? el[key] : null;
            };

            for (const wrapper of rechartsWrappers) {
                const fiber = getReactFiber(wrapper);
                if (!fiber) continue;

                let node = fiber;
                for (let i = 0; i < 50 && node; i++) {
                    const props = node.memoizedProps;
                    if (props && props.widgetData && Array.isArray(props.widgetData)) {
                        results.push(props.widgetData);
                        break;
                    }
                    node = node.return;
                }
            }

            return results;
        })()
    """)

    if not raw_charts:
        print("  WARNING: No chart data found in Recharts fiber tree")
        return []

    # Merge data from both charts (Chart 1: Case Status, Chart 2: Compressor Status)
    # Both share the same timestamps
    merged = {}
    for chart_data in raw_charts:
        for point in chart_data:
            ts = point.get("timestamp")
            if ts is None:
                continue
            if ts not in merged:
                merged[ts] = {"timestamp": ts}
            # Map chart fields to our standard names
            if "Control Temperature" in point and point["Control Temperature"] is not None:
                merged[ts]["control_temp"] = point["Control Temperature"]
            if "Defrost Terminate" in point and point["Defrost Terminate"] is not None:
                merged[ts]["defrost_terminate"] = point["Defrost Terminate"]
            if "Compressor Discharge Temp" in point and point["Compressor Discharge Temp"] is not None:
                merged[ts]["comp_discharge_temp"] = point["Compressor Discharge Temp"]
            if "Defrost Status" in point and point["Defrost Status"] is not None:
                merged[ts]["defrost_status"] = point["Defrost Status"]
            if "Refrigeration DO" in point and point["Refrigeration DO"] is not None:
                merged[ts]["refrigeration_do"] = point["Refrigeration DO"]
            if "Alarm" in point and point["Alarm"] is not None:
                merged[ts]["alarm"] = point["Alarm"]

    # Sort by timestamp and convert epoch ms to ISO string
    timeseries = sorted(merged.values(), key=lambda x: x["timestamp"])
    for point in timeseries:
        epoch_ms = point["timestamp"]
        point["timestamp_iso"] = datetime.fromtimestamp(epoch_ms / 1000, tz=None).strftime("%Y-%m-%d %H:%M:%S")

    print(f"  Extracted {len(timeseries)} time-series data points from charts")
    return timeseries


def calculate_stats(timeseries_data):
    """
    Calculate max, min, max_avg (avg of peaks), min_avg (avg of valleys),
    and cycles/hour from time-series data.
    """
    if not timeseries_data or len(timeseries_data) < 3:
        return {}

    stats = {}

    for field in ['control_temp', 'defrost_terminate', 'comp_discharge_temp']:
        values = [d.get(field) for d in timeseries_data if d.get(field) is not None]
        if not values:
            continue

        arr = np.array(values, dtype=float)
        stats[f'{field}_max'] = float(np.max(arr))
        stats[f'{field}_min'] = float(np.min(arr))

        # Find peaks and valleys for max_avg and min_avg
        peaks = []
        valleys = []
        for i in range(1, len(arr) - 1):
            if arr[i] > arr[i-1] and arr[i] > arr[i+1]:
                peaks.append(arr[i])
            elif arr[i] < arr[i-1] and arr[i] < arr[i+1]:
                valleys.append(arr[i])

        stats[f'{field}_max_avg'] = float(np.mean(peaks)) if peaks else stats[f'{field}_max']
        stats[f'{field}_min_avg'] = float(np.mean(valleys)) if valleys else stats[f'{field}_min']

    # Refrigeration cycles per hour
    # Use refrigeration_do (compressor on/off) if available, otherwise use control temp
    timestamps = [d.get('timestamp') for d in timeseries_data if d.get('timestamp') is not None]

    # Get compressor on/off signal
    compressor_states = [(d['timestamp'], d['refrigeration_do'])
                         for d in timeseries_data
                         if d.get('refrigeration_do') is not None and d.get('timestamp') is not None]

    if len(compressor_states) > 2:
        # Count transitions from OFF to ON (each = one cycle start)
        transitions = []
        for i in range(1, len(compressor_states)):
            prev_on = bool(compressor_states[i-1][1])
            curr_on = bool(compressor_states[i][1])
            if not prev_on and curr_on:  # OFF -> ON = cycle start
                transitions.append(compressor_states[i][0])

        total_hours = (compressor_states[-1][0] - compressor_states[0][0]) / 1000 / 3600
        if total_hours > 0 and transitions:
            overall_cph = len(transitions) / total_hours
            stats['cycles_per_hour_avg'] = round(overall_cph, 2)

            # Calculate per-hour buckets for max/min
            hourly_counts = {}
            for ts in transitions:
                hour_bucket = int(ts / 1000 / 3600)  # hour index
                hourly_counts[hour_bucket] = hourly_counts.get(hour_bucket, 0) + 1

            if hourly_counts:
                stats['cycles_per_hour_max'] = max(hourly_counts.values())
                stats['cycles_per_hour_min'] = min(hourly_counts.values())
    else:
        # Fallback: use control temp zero-crossings
        ctrl_temps = [d.get('control_temp') for d in timeseries_data if d.get('control_temp') is not None]
        if ctrl_temps and len(ctrl_temps) > 2 and len(timestamps) >= 2:
            arr = np.array(ctrl_temps, dtype=float)
            mean_val = np.mean(arr)
            crossings = np.where(np.diff(np.sign(arr - mean_val)))[0]
            num_cycles = len(crossings) / 2
            total_ms = timestamps[-1] - timestamps[0]
            total_hours = total_ms / 1000 / 3600
            if total_hours > 0:
                stats['cycles_per_hour_avg'] = round(num_cycles / total_hours, 2)

    return stats


def write_excel(day_dir, all_readings, all_stats, all_timeseries):
    """Write the collected data to Excel files."""
    wb = Workbook()

    # Sheet 1: Latest Readings
    ws1 = wb.active
    ws1.title = "Latest Readings"
    headers1 = [
        "Asset ID", "Store", "Model", "Module", "Timestamp",
        "Setpoint °F", "Ctrl Temp °F", "Defrost Term °F", "Comp Line Temp °F",
        "Evap Fan", "Compressor", "Defrost", "Light",
        "Alarm", "Low Alarm", "High Alarm"
    ]
    ws1.append(headers1)
    for r in all_readings:
        ws1.append([r.get(h, "") for h in [
            "asset_id", "store", "name", "module", "timestamp",
            "control_setpoint", "control_temp", "defrost_terminate", "compressor_line_temp",
            "evap_fan", "compressor_status", "defrost_status", "light_status",
            "alarm", "low_temp_alarm", "high_temp_alarm"
        ]])

    # Sheet 2: 24hr Stats
    ws2 = wb.create_sheet("24hr Stats")
    headers2 = [
        "Asset ID", "Store", "Model", "Module",
        "Ctrl Temp Max", "Ctrl Temp Min", "Ctrl Temp Max Avg", "Ctrl Temp Min Avg",
        "Defrost Term Max", "Defrost Term Min", "Defrost Term Max Avg", "Defrost Term Min Avg",
        "Comp Discharge Max", "Comp Discharge Min", "Comp Discharge Max Avg", "Comp Discharge Min Avg",
        "Cycles/Hr Max", "Cycles/Hr Min", "Cycles/Hr Avg"
    ]
    ws2.append(headers2)
    for s in all_stats:
        ws2.append([s.get(h, "") for h in [
            "asset_id", "store", "name", "module",
            "control_temp_max", "control_temp_min", "control_temp_max_avg", "control_temp_min_avg",
            "defrost_terminate_max", "defrost_terminate_min", "defrost_terminate_max_avg", "defrost_terminate_min_avg",
            "comp_discharge_temp_max", "comp_discharge_temp_min", "comp_discharge_temp_max_avg", "comp_discharge_temp_min_avg",
            "cycles_per_hour_max", "cycles_per_hour_min", "cycles_per_hour_avg"
        ]])

    summary_path = day_dir / "summary.xlsx"
    wb.save(summary_path)
    print(f"\nSaved summary to: {summary_path}")

    # Time series files per store
    stores = {}
    for ts in all_timeseries:
        store = ts.get("store", "unknown")
        if store not in stores:
            stores[store] = []
        stores[store].append(ts)

    for store_name, records in stores.items():
        ts_wb = Workbook()
        ts_ws = ts_wb.active
        ts_ws.title = "Time Series"
        ts_ws.append(["Asset ID", "Module", "Timestamp", "Control Temp °F", "Defrost Term °F", "Comp Discharge Temp °F"])
        for rec in records:
            ts_ws.append([
                rec.get("asset_id", ""),
                rec.get("module", ""),
                rec.get("timestamp", ""),
                rec.get("control_temp", ""),
                rec.get("defrost_terminate", ""),
                rec.get("comp_discharge_temp", "")
            ])
        safe_name = re.sub(r'[^\w\-]', '_', store_name)
        ts_path = day_dir / "timeseries" / f"{safe_name}.xlsx"
        ts_wb.save(ts_path)
        print(f"Saved time series for {store_name} to: {ts_path}")


def scrape_asset(page, asset_entry, day_dir, token):
    """Scrape all data for a single asset.
    asset_entry is the full dict from the asset list API.
    """
    asset_id = asset_entry["id"]
    print(f"\n--- Scraping asset {asset_id} ---")

    # Extract info from the asset list entry first (always available)
    serial = asset_entry.get("serialNumber", "")
    model_obj = asset_entry.get("model")
    model_name = model_obj.get("name", "") if isinstance(model_obj, dict) else ""
    ca = asset_entry.get("customAttributes") or {}
    site = ca.get("site_number", "")
    city = ca.get("city", "")

    # If asset list entry is missing info, try the detail API as fallback
    if not serial or not model_name:
        detail = get_asset_detail(page, asset_id, token)
        if detail and not isinstance(detail.get("message"), str):
            serial = serial or detail.get("serialNumber", "")
            if not model_name and detail.get("model"):
                model_name = detail["model"].get("name", "")
            if not site and detail.get("customAttributes"):
                ca = detail["customAttributes"]
                site = ca.get("site_number", "")
                city = ca.get("city", "")

    asset_info = {
        "asset_id": serial if serial else str(asset_id),
        "store": f"{site} - {city}".upper() if site else city,
        "name": model_name,
    }
    print(f"  Serial: {asset_info['asset_id']}, Model: {asset_info['name']}, Store: {asset_info['store']}")

    # Navigate to dashboard
    url = f"{BASE_URL}/tenants/{TENANT_ID}/features/asset-optimizer/assets/all-assets/details/{asset_id}/dashboard"
    page.goto(url)
    page.wait_for_load_state("domcontentloaded", timeout=30000)
    time.sleep(8)  # Wait for SPA to render dashboard and graphs

    # Get available modules
    modules = get_module_options(page)
    print(f"  Modules found: {modules}")

    all_readings = []
    all_stats = []
    all_timeseries = []

    for module_name in modules:
        print(f"  Selecting module: {module_name}")

        # Always select the module to ensure dashboard loads
        select_module(page, module_name)
        time.sleep(3)

        # Take screenshot
        safe_module = re.sub(r'[^\w\-]', '_', module_name)
        screenshot_path = day_dir / "screenshots" / f"{asset_info['asset_id']}_{safe_module}.png"
        page.screenshot(path=str(screenshot_path), full_page=True)
        print(f"  Screenshot saved: {screenshot_path}")

        # Extract dashboard values from DOM
        dashboard_data = extract_dashboard_data(page)
        status_data = extract_status_tiles(page)

        reading = {
            **asset_info,
            "module": module_name,
            "timestamp": dashboard_data.get("last_data_timestamp", ""),
            "control_setpoint": dashboard_data.get("control_setpoint", ""),
            "control_temp": dashboard_data.get("control_temp", ""),
            "defrost_terminate": dashboard_data.get("defrost_terminate", ""),
            "compressor_line_temp": dashboard_data.get("compressor_line_temp", ""),
            "evap_fan": status_data.get("Evaporator Fan", ""),
            "compressor_status": status_data.get("Compressor Status", ""),
            "defrost_status": status_data.get("Defrost Status", ""),
            "light_status": status_data.get("Light Status", ""),
            "alarm": status_data.get("Alarm", ""),
            "low_temp_alarm": status_data.get("Low Temperature Alarm", ""),
            "high_temp_alarm": status_data.get("High Temperature Alarm", ""),
        }
        all_readings.append(reading)
        print(f"  Readings: Setpoint={reading['control_setpoint']}, "
              f"Ctrl={reading['control_temp']}, "
              f"Defrost={reading['defrost_terminate']}, "
              f"CompLine={reading['compressor_line_temp']}")

        # Extract 24hr time-series data from Recharts charts
        timeseries = extract_timeseries_from_charts(page)

        # Calculate stats from time-series data
        stats = calculate_stats(timeseries)
        stat_row = {
            **asset_info,
            "module": module_name,
            **stats,
        }
        all_stats.append(stat_row)

        # Add store/asset info to each timeseries point for Excel output
        for point in timeseries:
            all_timeseries.append({
                **asset_info,
                "module": module_name,
                "timestamp": point.get("timestamp_iso", ""),
                "control_temp": point.get("control_temp", ""),
                "defrost_terminate": point.get("defrost_terminate", ""),
                "comp_discharge_temp": point.get("comp_discharge_temp", ""),
            })

    return all_readings, all_stats, all_timeseries


def main():
    print("=" * 60)
    print("OneConnect StoreConnect Pulse - Asset Dashboard Scraper")
    print("=" * 60)

    day_dir = setup_daily_folder()
    print(f"Output folder: {day_dir}")

    with sync_playwright() as p:
        # Always launch a visible browser
        print("\nLaunching browser...")
        browser = p.chromium.launch(headless=False, slow_mo=300)

        # Try to use saved session, otherwise fresh login
        if AUTH_STATE_FILE.exists():
            context = browser.new_context(
                storage_state=str(AUTH_STATE_FILE),
                viewport={"width": 1536, "height": 900}
            )
        else:
            context = browser.new_context(viewport={"width": 1536, "height": 900})

        page = context.new_page()

        # Navigate to asset list to check if logged in
        page.goto(f"{BASE_URL}/tenants/{TENANT_ID}/features/asset-optimizer/assets/all-assets/list")
        page.wait_for_load_state("domcontentloaded", timeout=30000)
        time.sleep(5)

        print(f"Current URL after navigation: {page.url}")

        # Check if we need to log in — also check if token exists
        token_check = page.evaluate("localStorage.getItem('TOKEN')")
        needs_login = "/login" in page.url or "/redirecter" in page.url or not token_check
        print(f"Token found: {bool(token_check)}, Needs login: {needs_login}")

        if needs_login:
            print("\n" + "!" * 60)
            print("  ACTION NEEDED: Log in to the browser window that opened!")
            print("  Enter your email and password like you normally do.")
            print("  The script will wait and continue automatically.")
            print("!" * 60 + "\n")

            # If we're not already on the login page, go there
            if "/login" not in page.url:
                page.goto(f"{BASE_URL}/login")

            # Wait for login to complete (token appears in localStorage)
            for i in range(600):  # 10 minutes
                time.sleep(1)
                try:
                    # Try clicking "Yes" on "stay signed in" prompt if it appears
                    stay_btn = page.locator('text=/Yes/i')
                    if stay_btn.count() > 0 and stay_btn.first.is_visible():
                        print("  Clicking 'Stay signed in'...")
                        stay_btn.first.click()
                        time.sleep(3)
                except Exception:
                    pass

                try:
                    token = page.evaluate("localStorage.getItem('TOKEN')")
                    if token and token.startswith("Bearer"):
                        break
                except Exception:
                    pass

                if i % 30 == 0 and i > 0:
                    print(f"  Still waiting for login... ({i}s)")
            else:
                print("ERROR: Login timed out.")
                browser.close()
                return

            time.sleep(5)
            print("Login detected!")

            # Navigate to asset list after login
            page.goto(f"{BASE_URL}/tenants/{TENANT_ID}/features/asset-optimizer/assets/all-assets/list")
            page.wait_for_load_state("domcontentloaded", timeout=30000)
            time.sleep(5)

        # Save session for next time
        context.storage_state(path=str(AUTH_STATE_FILE))

        # Get auth token
        token = page.evaluate("localStorage.getItem('TOKEN')")
        if not token:
            print("ERROR: Could not get auth token. Please check your login.")
            browser.close()
            return

        print(f"Authenticated. Token starts with: {token[:20]}...")

        # Fetch asset IDs from API
        print(f"\nFetching asset list (first {NUM_ASSETS_TO_SCRAPE} assets)...")
        asset_list_json = page.evaluate(f"""
            fetch('{API_BASE}/tenants/{TENANT_ID}/tenant-api/asset-optimizer/v1/assets?pageNumber=0&pageSize={NUM_ASSETS_TO_SCRAPE}', {{
                headers: {{'Authorization': '{token}'}}
            }}).then(r => r.text())
        """)
        try:
            asset_list = json.loads(asset_list_json)
            # API returns {data: [...], meta: {...}}
            if isinstance(asset_list, dict):
                items = asset_list.get("data") or asset_list.get("content") or asset_list.get("items") or []
            elif isinstance(asset_list, list):
                items = asset_list
            else:
                items = []

            if not items:
                print(f"  Could not find asset list. Using defaults.")
                items = [{"id": 8043}]

            # Keep the full asset entry — it has model, serial, store info
            assets = items[:NUM_ASSETS_TO_SCRAPE]
            print(f"Found {len(assets)} assets")
        except Exception as e:
            print(f"WARNING: Could not parse asset list ({e}), using default")
            assets = [{"id": 8043}]

        # Scrape assets
        all_readings = []
        all_stats = []
        all_timeseries = []

        for asset_entry in assets:
            readings, stats, timeseries = scrape_asset(page, asset_entry, day_dir, token)
            all_readings.extend(readings)
            all_stats.extend(stats)
            all_timeseries.extend(timeseries)

        # Write results
        write_excel(day_dir, all_readings, all_stats, all_timeseries)

        print(f"\n{'=' * 60}")
        print(f"Done! Scraped {len(all_readings)} module readings from {len(asset_ids)} assets.")
        print(f"Results saved to: {day_dir}")
        print(f"{'=' * 60}")

        browser.close()


if __name__ == "__main__":
    main()
